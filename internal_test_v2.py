import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import os
import scipy.io as scio
import random
import pandas as pd
from openpyxl import load_workbook
import h5py
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy import stats

def set_seed(seed):
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def split():
    CT_value_file = "./data/CT_value_20231219"
    train = "train.txt"
    test = "test.txt"
    f1 = open(train, "w")
    f2 = open(test, "w")
    cross = 5
    file_list = os.listdir(CT_value_file)
    random.seed(123)
    random.shuffle(file_list)
    a = int(len(file_list) / cross)
    train_list = []
    test_list = []
    for q in range(len(file_list)):
        if q < a:
            f2.writelines(file_list[q] + "\n")
            test_list.append(file_list[q])
        else:
            f1.writelines(file_list[q] + "\n")
            train_list.append(file_list[q])
    f1.close()
    f2.close()

def total_data():
    CT_value_file = "./data/CT_value_20231219"
    file_list = os.listdir(CT_value_file)
    treatment_file = "./data/treatment_para/"
    total1=0
    total2=0
    sub_data = {}
    for file in file_list:
        path = os.path.join(treatment_file, file)+".csv"
        data = pd.read_csv(path)
        sub_data["Flag"] = data.iloc[:, 19]
        aa=data.iloc[:, 19]
        total1 += 1
        total2+=len(aa)
    print("patients number:%.2f"%total1)
    print("treatment number:%.2f"%total2)

def actual_data():
    CT_value_file = "./data/CT_value_20231219"
    file_list = os.listdir(CT_value_file)

    treatment_file = "./data/treatment_para/"
    patient=0
    total=0
    sub_data = {}
    for file in file_list:
        path = os.path.join(treatment_file, file)+".csv"
        data = pd.read_csv(path)
        sub_data["Flag"] = data.iloc[:, 19]
        aa=data.iloc[:, 19]
        # print(aa[1])
        indices_of_no = aa[aa =='No             '].index
        if len(indices_of_no)==len(aa):
            patient+=1
        total+=len(indices_of_no)
    print("effective patients:%.2f"%patient)
    print("effective number:%.2f"%total)

class AttentionLayer(nn.Module):
    def __init__(self, hidden_dim):
        super(AttentionLayer, self).__init__()
        self.W = nn.Linear(hidden_dim,hidden_dim)
        self.V = nn.Linear(hidden_dim, 1)
        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        # weighted_x = x+x * attn_weights  # Weighted sum of features

        attn_weights = self.dropout(attn_weights)
        weighted_x = self.layer_norm(x + x * attn_weights)
        return weighted_x  # Sum across features

class FactorAttentionLayer(nn.Module):
    def __init__(self, hidden_dim):
        super(FactorAttentionLayer, self).__init__()
        self.W = nn.Linear(hidden_dim, hidden_dim)
        self.V = nn.Linear(hidden_dim, 1)
        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        # weighted_x = x + x * attn_weights  # Weighted sum of features

        attn_weights = self.dropout(attn_weights)
        weighted_x = self.layer_norm(x + x * attn_weights)

        return weighted_x  # Sum across features

class MLPModel7(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel7, self).__init__()

        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        # self.factor_attention2 = FactorAttentionLayer(input_dim)
        # self.factor_attention = MultiHeadFactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, inputs):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        # 通过对应的线性层处理每个输入特征
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs)]
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
        # x_all=x_all.reshape(x_all.shape[0], -1)

        # 应用注意力机制
        xx = self.factor_attention(x_all.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#attention2；+dropout layernorm+5.455
        # xx = self.attention(x_all).reshape(x_all.shape[0], -1)#attention3

        # xx = self.factor_attention(x_all.permute(0, 2, 1))
        # xx = self.attention(xx.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#mutil-attention

        # 应用残差连接
        x = self.fc_hidden(xx)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

class MLPModel8(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel8, self).__init__()
        self.fc=nn.Linear(1024, hidden_dim)
        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        # self.factor_attention2 = FactorAttentionLayer(input_dim)
        # self.factor_attention = MultiHeadFactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, inputs):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        fc = self.relu(self.fc(inputs[0]))
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs[1:])]
        processed_inputs=[fc]+processed_inputs
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
        # x_all=x_all.reshape(x_all.shape[0], -1)

        # 应用注意力机制
        xx = self.factor_attention(x_all.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#attention2；+dropout layernorm+5.455
        # xx = self.attention(x_all).reshape(x_all.shape[0], -1)#attention3

        # xx = self.factor_attention(x_all.permute(0, 2, 1))
        # xx = self.attention(xx.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#mutil-attention

        # 应用残差连接
        x = self.fc_hidden(xx)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

# class mm(nn.Module):
#     def __init__(self, input_dim, hidden_dim, output_dim):
#         super(mm, self).__init__()
#
#         self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
#         # self.attention = AttentionLayer(hidden_dim)
#         self.factor_attention = FactorAttentionLayer(input_dim)
#         # self.factor_attention2 = FactorAttentionLayer(input_dim)
#         # self.factor_attention = MultiHeadFactorAttentionLayer(input_dim)
#         self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
#         self.fc_output = nn.Linear(hidden_dim, output_dim)
#         self.relu = nn.ReLU()
#         self.dropout = nn.Dropout(0.2)
#
#     def forward(self, inputs):
#         # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
#         # 通过对应的线性层处理每个输入特征
#         processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs)]
#         x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
#         # x_all=x_all.reshape(x_all.shape[0], -1)
#
#         # 应用注意力机制
#         xx = self.factor_attention(x_all.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#attention2；+dropout layernorm+5.455
#         # xx = self.attention(x_all).reshape(x_all.shape[0], -1)#attention3
#
#         # xx = self.factor_attention(x_all.permute(0, 2, 1))
#         # xx = self.attention(xx.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#mutil-attention
#
#         # 应用残差连接
#         x = self.fc_hidden(xx)
#         x = self.relu(x)
#         x = self.dropout(x)
#
#         # 最终输出
#         x = self.fc_output(x)
#
#         return x

if __name__ == '__main__':
    set_seed(64)  # 可以选择任何整数值作为随机种子
    cross=5
    models = ["model7_dis_dur_energy_area_size_sdr_avp_seq","model8_ang_dis_dur_energy_area_size_sdr_avp_seq"]
    for j in range(len(models)):
        mm = models[j]
        # if j<1:
        #     continue
        for i in range(cross):
            # if i < 4:
            #     continue
            fold = "./fold_" + str(i + 1)
            os.makedirs(fold, exist_ok=True)
            model_path = fold + "/" + mm
            # latest_model = model_path + '/latest_model.pth'
            best_model = model_path + '/best_model.pth'
            in_results = fold + '/internal_results/'
            if not os.path.exists(in_results):
                os.mkdir(in_results)
            test_record = in_results + mm + ".txt"
            loss_record = in_results + mm + "_loss.txt"
            # ex_results = fold + '/external_results/'
            # 读取数据
            with open(f"./flist/test_fold_{i + 1}.txt", "r") as f_test:
                f_test_list = f_test.read().split("\n")

            test_list = []
            for file in f_test_list:
                if file:  # 检查是否为空字符串
                    test_list.append(file.split("\n")[0])

            model = torch.load(best_model)  #
            criterion = nn.L1Loss(reduction="none")
            model.eval()
            with torch.no_grad():
                # 使用训练好的模型进行预测
                # 假设 test_data 是测试数据
                f_test = open(test_record, "w")
                loss_test = open(loss_record, "w")

                CT_value_file = "./data/CT_value_20231219"
                CT_value = {}
                for file in test_list:
                    ID = file.split('.')[0]
                    k = 0
                    path = os.path.join(CT_value_file, file)
                    with open(path, 'r', encoding='UTF-8') as f:
                        sub_data = []
                        for line in f.readlines():
                            k = k + 1
                            if k < 4:
                                continue
                            line = line.strip('\n')  # 去掉换行符\n
                            data_list = line.split('\t')[1:]
                            data_list = list(map(float, data_list))  # 将每一行以空格为分隔符转换成列表
                            mea = sum(data_list) / len(data_list)
                            sub_data.append(mea)  # 全部除以1000
                        sub_data = np.array(sub_data)
                        CT_value[ID] = sub_data

                    # SDR score of patient;patients x 1
                    path = "./data/CT_proj_data20231216.xlsx"
                    look_up_table_row_start = 2
                    look_up_table_row_number = 90
                    skull_area = {}
                    skull_score = {}
                    measurement = {}
                    ages = {}
                    disea = {}
                    look_up_table_excel = load_workbook(path)
                    look_up_table_all_sheet = look_up_table_excel.sheetnames
                    look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
                    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
                        ID = look_up_table_sheet.cell(i, 1).value
                        sub_data1 = look_up_table_sheet.cell(i, 8).value
                        sub_data2 = look_up_table_sheet.cell(i, 11).value
                        sub_data3 = look_up_table_sheet.cell(i, 17).value
                        sub_data4 = look_up_table_sheet.cell(i, 20).value
                        sub_data5 = look_up_table_sheet.cell(i, 22).value
                        ages[ID] = sub_data1
                        disea[ID] = sub_data2
                        skull_area[ID] = sub_data3
                        skull_score[ID] = sub_data4
                        measurement[ID] = sub_data5

                total_loss = []
                Data_ver = "./data/Data_ver5"
                for file in test_list:
                    CT_density = CT_value[file]
                    Brain_area = skull_area[file]
                    age = ages[file]
                    dis = disea[file]
                    total_sdr = skull_score[file]
                    Brain_size = measurement[file]
                    path = os.path.join(Data_ver, file)
                    DATA_Mat = scio.loadmat(path)
                    # print(DATA_Mat.keys())
                    AvgPower = DATA_Mat['Elements']['AvgPower'][0][0][0, :]
                    Thickness = DATA_Mat['Elements']['Thickness'][0][0][0, :]
                    SDR = DATA_Mat['Elements']['SDR'][0][0][0, :]
                    Angle = DATA_Mat['Elements']['Angle'][0][0][0, :]
                    ElementPosition = DATA_Mat['Elements']['ElementPosition'][0][0][0, :]
                    OnOff = DATA_Mat['Elements']['OnOff'][0][0][0, :]

                    # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                    # ActualEnergy = DATA_Mat['Sonication']['ActualEnergy'][0][0][:, 0]

                    ActualDuration = DATA_Mat['Sonication']['PlannedDuration'][0][0][:, 0]  # Sonication.PlannedDuration
                    ActualEnergy = DATA_Mat['Sonication']['PlannedEnergy'][0][0][:, 0]
                    MaxAvgT = DATA_Mat['Sonication']['MaxAvgT'][0][0][:, 0]  # max average temperature which our target.
                    # sub_list = list(range(len(AvgPower)))

                    treatment_file = "./data/treatment_para/"
                    path = os.path.join(treatment_file, file) + ".csv"
                    data = pd.read_csv(path)
                    aa = data.iloc[:, 19]
                    # print(aa[1])
                    sub_list = aa[aa == 'No             '].index

                    AvgPower_in = []
                    Thickness_in = []
                    SDR_in = []
                    Angle_in = []
                    ElementPosition_in = []
                    # ActualEnergy_in=[]

                    MaxAvgT_out = []
                    for m in sub_list:
                        AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                        Thickness_in.append(Thickness[m][:, 0] * CT_density * OnOff[m][:, 0])
                        SDR_in.append(SDR[m][:, 0] * OnOff[m][:, 0])  # 是否受能量启动影像
                        Angle_in.append(Angle[m][:, 0] * OnOff[m][:, 0])
                        ElementPosition_in.append(ElementPosition[m][:, :] * OnOff[m][:, 0][:, np.newaxis].
                                                  repeat(ElementPosition[m][:, :].shape[1], axis=1))
                    AvgPower_in = np.array(AvgPower_in)
                    Thickness_in = np.array(Thickness_in) / 20000
                    SDR_in = np.array(SDR_in)
                    SDR_in = -np.nan_to_num(SDR_in)
                    Angle_in = np.array(Angle_in) / 30
                    ElementPosition_in = np.array(ElementPosition_in) / 200

                    ActualDuration_in = ActualDuration[sub_list] / 10
                    ActualEnergy_in = ActualEnergy[sub_list] / 10000
                    age_in = np.full_like(ActualDuration_in, 1) * age / 100
                    dis_in = np.full_like(ActualDuration_in, 1) * dis / 40

                    Brain_area_in = np.full_like(ActualDuration_in, 1) * Brain_area / 400
                    Brain_size_in = np.full_like(ActualDuration_in, 1) * Brain_size / 60
                    total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr
                    y_train0 = MaxAvgT[sub_list]
                    X_traina = AvgPower_in
                    X_traina = np.mean(X_traina, axis=1).reshape(-1, 1)
                    X_trainb = Thickness_in
                    X_trainb = np.mean(X_trainb, axis=1).reshape(-1, 1)

                    X_trainc = Angle_in#age_in[:, np.newaxis]
                    X_traind = dis_in[:, np.newaxis]
                    X_traine = ActualDuration_in[:, np.newaxis]
                    X_trainf = ActualEnergy_in[:, np.newaxis]

                    X_traing = Brain_area_in[:, np.newaxis]
                    X_trainh = Brain_size_in[:, np.newaxis]
                    X_traini = total_sdr_in[:, np.newaxis]
                    x_trainz = np.array(sub_list)[:, np.newaxis] / 20

                    y_train0 = y_train0[:, np.newaxis]

                    # 前向传播
                    # 将训练数据转换为 PyTorch 的 Tensor 类型
                    X_traina = torch.from_numpy(X_traina).float()
                    X_trainb = torch.from_numpy(X_trainb).float()
                    X_trainc = torch.from_numpy(X_trainc).float()
                    X_traind = torch.from_numpy(X_traind).float()
                    X_traine = torch.from_numpy(X_traine).float()
                    X_trainf = torch.from_numpy(X_trainf).float()
                    X_traing = torch.from_numpy(X_traing).float()
                    X_trainh = torch.from_numpy(X_trainh).float()
                    X_traini = torch.from_numpy(X_traini).float()
                    x_trainz = torch.from_numpy(x_trainz).float()

                    y_train = torch.from_numpy(y_train0).float()
                    X_m1_list = [X_traine, X_trainf, X_traini, X_traina, X_trainh, X_traing, X_traind, x_trainz]
                    # X_m1_list = [X_trainc, X_traine, X_trainf, X_traini, X_traina, X_trainh, X_traing, X_traind,x_trainz]
                    outputs = model(X_m1_list)
                    absolute_errors = criterion(outputs, y_train)
                    absolute_errors = np.array(absolute_errors)[:, 0]
                    # absolute_errors = torch.abs(outputs[:,0], y_train[:,0])
                    mean_absolute_error = absolute_errors.mean()
                    total_loss.append(mean_absolute_error)
                    sample_variance = absolute_errors.var()
                    loss = str(mean_absolute_error)  #
                    predict = outputs.detach().numpy()
                    out = np.hstack((y_train0, predict))  # target+predict
                    # f_test.writelines(out)
                    for ii in range(len(out)):
                        f_test.write(str(out[ii]) + '\n')
                    f_test.writelines("\n")
                    loss_test.write(str(loss) + '\n')
                errors = np.array([total_loss])
                mean_error = np.mean(errors)# 计算均值和标准差
                std_error = np.std(errors, ddof=1)
                n = len(errors)# 样本量
                se = std_error / np.sqrt(n)# 计算标准误
                # 确定z值（95%置信水平）
                z_value = stats.norm.ppf(0.975)  # 对应95%置信水平的z值
                lower_bound = mean_error - z_value * se# 计算置信区间
                upper_bound = mean_error + z_value * se
                # 将均值和置信区间追加写入 f_test 文件
                f_test.write(f"Mean Error: {mean_error:.4f}\n")
                f_test.write(f"95% Confidence Interval: [{lower_bound:.4f}, {upper_bound:.4f}]\n")
                f_test.write("\n")  # 添加一个空行分隔不同的结果块