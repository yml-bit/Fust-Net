import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import os
import scipy.io as scio
import random
import pandas as pd
from openpyxl import load_workbook
import h5py
import scipy

def split():
    CT_value_file = "./data/CT_value_20231219"
    train = "train.txt"
    test = "test.txt"
    f1 = open(train, "w")
    f2 = open(test, "w")
    cross = 5
    file_list = os.listdir(CT_value_file)
    random.seed(123)
    random.shuffle(file_list)
    a = int(len(file_list) / cross)
    train_list = []
    test_list = []
    for q in range(len(file_list)):
        if q < a:
            f2.writelines(file_list[q] + "\n")
            test_list.append(file_list[q])
        else:
            f1.writelines(file_list[q] + "\n")
            train_list.append(file_list[q])
    f1.close()
    f2.close()

def total_data():
    CT_value_file = "./data/CT_value_20231219"
    file_list = os.listdir(CT_value_file)
    treatment_file = "./data/treatment_para/"
    total1=0
    total2=0
    sub_data = {}
    for file in file_list:
        path = os.path.join(treatment_file, file)+".csv"
        data = pd.read_csv(path)
        sub_data["Flag"] = data.iloc[:, 19]
        aa=data.iloc[:, 19]
        total1 += 1
        total2+=len(aa)
    print("patients number:%.2f"%total1)
    print("treatment number:%.2f"%total2)

def actual_data():
    CT_value_file = "./data/CT_value_20231219"
    file_list = os.listdir(CT_value_file)

    treatment_file = "./data/treatment_para/"
    patient=0
    total=0
    sub_data = {}
    for file in file_list:
        path = os.path.join(treatment_file, file)+".csv"
        data = pd.read_csv(path)
        sub_data["Flag"] = data.iloc[:, 19]
        aa=data.iloc[:, 19]
        # print(aa[1])
        indices_of_no = aa[aa =='No             '].index
        if len(indices_of_no)==len(aa):
            patient+=1
        total+=len(indices_of_no)
    print("effective patients:%.2f"%patient)
    print("effective number:%.2f"%total)

#model1_dur  #val lr=1e-3:21.21
#model2_dur_energy  #val lr=1e-3:10.22
#model3_dur_energy_sdr  #lr=1e-3:6.92
#model3_dur_energy_area  #lr=1e-3:10.2
#model3_dur_energy_size  #lr=1e-3:9.93
#model4_dur_energy_area_sdr  #lr=1e-3:7.08
#model4_dur_energy_size_sdr  #lr=1e-3:#6.676
#model5_dur_energy_area_size_sdr  #lr=1e-3:6.525
#model6_dis_dur_energy_area_size_sdr  #lr=1e-3:6.069
#model6_age_dur_energy_area_size_sdr  #lr=1e-3:6.38
#model7_dis_age_dur_energy_area_size_sdr  #lr=1e-3:6.5090
#model8_age_dur_energy_area_size_sdr_thick  #lr=1e-3:6.5558

# X_traina = AvgPower_in
# X_trainb = Thickness_in
# X_trainc = age_in[:, np.newaxis]
# X_traind = dis_in[:, np.newaxis]   nn
# X_traine = ActualDuration_in[:, np.newaxis]  n
# X_trainf = ActualEnergy_in[:, np.newaxis]   nn
# X_traing = Brain_area_in[:, np.newaxis]  n
# X_trainh = Brain_size_in[:, np.newaxis]  n
# X_traini = total_sdr_in[:, np.newaxis]   nn
class MLPModel1(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel1, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(64, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))

        x = self.relu(self.fc1(x_E1))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel2(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel2, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)

        self.fc1 = nn.Linear(128, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))


        x_all = torch.cat((x_E1,x_F1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel3(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel3, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(192, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(H))


        x_all = torch.cat((x_E1,x_F1,x_G1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel4(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel4, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(256, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(H))
        x_H1 = self.relu(self.fc_H1(I))


        x_all = torch.cat((x_E1,x_F1,x_G1,x_H1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel5(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel5, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(320, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel6(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel6, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        # x_C1 = self.relu(self.fc_C1(E))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_D1,x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel66(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel66, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        # x_C1 = self.relu(self.fc_C1(E))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_D1,x_E1*0,x_F1,x_G1*0,x_H1*0,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x
class MLPModel7(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel7, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(448, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_C1 = self.relu(self.fc_C1(C))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_C1,x_D1,x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel8(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel8, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        self.fc_B1 = nn.Linear(1024, 512)
        self.fc_B2 = nn.Linear(512, 64)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        # self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        x_B1 = self.relu(self.fc_B1(B))
        x_B2 = self.relu(self.fc_B2(x_B1))
        # x_C1 = self.relu(self.fc_C1(E))
        # x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_B2,x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel223(nn.Module):
    def __init__(self,input_dim, output_dim):
        super(MLPModel223, self).__init__()
        importance_D = 0.75
        importance_E = 0.25
        # 对A类数据的处理网络（使用1D卷积）
        self.fc_A1 = nn.Linear(1024, 512)
        self.fc_B1 = nn.Linear(1024, 512)
        self.fc_C1 = nn.Linear(1024, 256)
        # 对B类数据的处理网络（使用1D卷积）

        # 对C类数据的处理网络（使用1D卷积）
        self.fc_D1 = nn.Sequential(
            nn.Conv1d(4, 64, kernel_size=3, padding=1),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
            nn.Flatten(),
            nn.Linear(64 * (1024 // 2), 256)
        )
        # 对D和E类数据进行线性映射并调整权重
        self.fc_E = nn.Linear(1, 32)
        self.fc_F = nn.Linear(1, 32)

        # self.fc1 = nn.Linear(1600, 512)
        # self.fc2 = nn.Linear(512, 256)
        # self.fc3 = nn.Linear(256, output_dim)  # 隐藏层到输出层
        # self.relu = nn.ReLU()  # 非线性激活函数

        self.fc_output = nn.Sequential(
            nn.Linear(1600, 512),
            nn.ReLU(),
            nn.Linear(512, 1),
            nn.Sigmoid()  # 使用Sigmoid激活函数将输出范围压缩到(0, 1)之间，然后通过后处理将其转换为(30, 60)
        )
    def forward(self, A,B, C, D, E,F,G,H):
        x_A = self.fc_A1(A)  # 将A从batchx1024x2转为2x1024xbatch以适应卷积层
        x_B = self.fc_B1(B)  # 将B从batchx1024x1转为1x1024xbatch以适应卷积层
        x_C = self.fc_C1(C)  # 将C从batchx1024x4转为4x1024xbatch以适应卷积层
        x_D = self.fc_D1(D.permute(0, 2, 1))
        x_E = torch.relu(self.fc_E(E))
        x_F = torch.relu(self.fc_F(F))

        # 应用可学习的重要程度权重
        # x_D *= self.importance_D.unsqueeze(-1).unsqueeze(0).expand_as(x_D)
        # x_E *= self.importance_E.unsqueeze(-1).unsqueeze(0).expand_as(x_E)

        # 合并所有输入
        x_all= torch.cat((x_A, x_B, x_C, x_D,x_E,x_F), dim=-1)
        # x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        # output= self.fc3(x)
        output=self.fc_output(x_all)
        # 后处理将预测值映射到30-60范围内
        return ((output * (60 - 30)) + 30).squeeze(-1)

# X_traine = ActualDuration_in[:, np.newaxis]
# X_trainf = ActualEnergy_in[:, np.newaxis]
# X_traing = Brain_area_in[:, np.newaxis]
# X_trainh = Brain_size_in[:, np.newaxis]
# X_traini = total_sdr_in[:, np.newaxis]
# Press the green button in the gutter to run the script.

if __name__ == '__main__':
    #random split
    # split()#初次运行才需要执行
    total_data()
    actual_data()

    #读取划分好的文件
    with open("train.txt", "r") as tf:#记得检查txt文档最后一行是否为空
        first_list = tf.read().split("\n")
    train_list=[]
    for file in first_list:
        # if "PTT" in file:
        #     continue
        train_list.append(file.split("\n")[0])

    with open("test.txt", "r") as tf:
        second_list = tf.read().split("\n")
    test_list=[]
    for file in second_list:
        # if "PTT" in file:
        #     continue
        test_list.append(file.split("\n")[0])

    # 根据数据格式定义输入维度
    input_dim = 1024 * 3 + 2  # 超声波阵列位置坐标数 + 每个阵列点能量 + 球壳密度 + 球壳厚度
    output_dim = 1  # 输出预测的温度
    model = MLPModel66(input_dim, output_dim) # 实例化模型
    model_path = "./model66_dis_energy_sdr"
    if not os.path.exists(model_path):
        os.mkdir(model_path)
    latest_model=model_path+'/latest_model.pth'
    best_model=model_path+'/best_model.pth'

    optimizer = optim.Adam(model.parameters(), lr=3e-3)  # 1e-3 3e-3
    criterion=nn.MSELoss()
    # criterion = nn.L1Loss(reduction='mean') #nn.MSELoss()
    # Section "OutputClass"
    #     Identifier "nvidia"
    #     MatchDriver "nvidia-drm"
    #     Driver "nvidia"
    #     Option "AllowEmptyInitialConfiguration"
    #     ModulePath "/usr/lib/x86_64-linux-gnu/nvidia/xorg"
    # EndSection
    def val_test(test_list):
        model.eval()
        with torch.no_grad():
            # 使用训练好的模型进行预测
            # 假设 test_data 是测试数据
            test_record = "test_record.txt"
            f_test = open(test_record, "w")

            CT_value_file = "./data/CT_value_20231219"
            CT_value = {}
            for file in test_list:
                ID = file.split('.')[0]
                k = 0
                path = os.path.join(CT_value_file, file)
                with open(path, 'r', encoding='UTF-8') as f:
                    sub_data = []
                    for line in f.readlines():
                        k = k + 1
                        if k < 4:
                            continue
                        line = line.strip('\n')  # 去掉换行符\n
                        data_list = line.split('\t')[1:]
                        data_list = list(map(float, data_list))  # 将每一行以空格为分隔符转换成列表
                        mea = sum(data_list) / len(data_list)
                        sub_data.append(mea)  # 全部除以1000
                    sub_data = np.array(sub_data)
                    CT_value[ID] = sub_data

            # SDR score of patient;patients x 1
            path = "./data/CT_proj_data20231216.xlsx"
            look_up_table_row_start = 2
            look_up_table_row_number = 90
            skull_area = {}
            skull_score = {}
            measurement = {}
            ages = {}
            disea = {}
            look_up_table_excel = load_workbook(path)
            look_up_table_all_sheet = look_up_table_excel.sheetnames
            look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
            for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
                ID = look_up_table_sheet.cell(i, 1).value
                sub_data1 = look_up_table_sheet.cell(i, 8).value
                sub_data2 = look_up_table_sheet.cell(i, 11).value
                sub_data3 = look_up_table_sheet.cell(i, 17).value
                sub_data4 = look_up_table_sheet.cell(i, 20).value
                sub_data5 = look_up_table_sheet.cell(i, 22).value
                ages[ID] = sub_data1
                disea[ID] = sub_data2
                skull_area[ID] = sub_data3
                skull_score[ID] = sub_data4
                measurement[ID] = sub_data5

            total_loss = []
            Data_ver = "./data/Data_ver5"
            for file in test_list:
                CT_density = CT_value[file]
                Brain_area=skull_area[file]
                age=ages[file]
                dis=disea[file]
                total_sdr=skull_score[file]
                Brain_size=measurement[file]
                path = os.path.join(Data_ver, file)
                DATA_Mat = scio.loadmat(path)
                # print(DATA_Mat.keys())
                AvgPower = DATA_Mat['Elements']['AvgPower'][0][0][0, :]
                Thickness = DATA_Mat['Elements']['Thickness'][0][0][0, :]
                SDR = DATA_Mat['Elements']['SDR'][0][0][0, :]
                Angle = DATA_Mat['Elements']['Angle'][0][0][0, :]
                ElementPosition = DATA_Mat['Elements']['ElementPosition'][0][0][0, :]
                OnOff = DATA_Mat['Elements']['OnOff'][0][0][0, :]

                # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                # ActualEnergy = DATA_Mat['Sonication']['ActualEnergy'][0][0][:, 0]#Sonication.PlannedEnergy

                ActualDuration = DATA_Mat['Sonication']['PlannedDuration'][0][0][:, 0]#Sonication.PlannedDuration
                ActualEnergy = DATA_Mat['Sonication']['PlannedEnergy'][0][0][:, 0]
                MaxAvgT = DATA_Mat['Sonication']['MaxAvgT'][0][0][:, 0]  # max average temperature which our target.

                # sub_list = list(range(len(AvgPower)))
                treatment_file = "./data/treatment_para/"
                path = os.path.join(treatment_file, file) + ".csv"
                data = pd.read_csv(path)
                aa = data.iloc[:, 19]
                # print(aa[1])
                sub_list = aa[aa == 'No             '].index

                AvgPower_in = []
                Thickness_in = []
                SDR_in = []
                Angle_in=[]
                ElementPosition_in=[]
                # ActualEnergy_in=[]

                MaxAvgT_out = []
                for m in sub_list:
                    AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                    Thickness_in.append(Thickness[m][:, 0] * CT_density * OnOff[m][:, 0])
                    SDR_in.append(SDR[m][:, 0]* OnOff[m][:, 0])  # 是否受能量启动影像
                    Angle_in.append(Angle[m][:, 0]*OnOff[m][:,0])
                    ElementPosition_in.append(ElementPosition[m][:, :] * OnOff[m][:, 0][:, np.newaxis].
                                              repeat(ElementPosition[m][:, :].shape[1], axis=1))
                AvgPower_in = np.array(AvgPower_in)
                Thickness_in = np.array(Thickness_in)/1000
                SDR_in=np.array(SDR_in)
                SDR_in=-np.nan_to_num(SDR_in)
                Angle_in=np.array(Angle_in)/30
                ElementPosition_in = np.array(ElementPosition_in)/200

                ActualDuration_in = ActualDuration[sub_list]/10
                ActualEnergy_in = ActualEnergy[sub_list]/10000
                age_in = np.full_like (ActualDuration_in, 1)*age/100
                dis_in = np.full_like (ActualDuration_in, 1)*dis/40

                Brain_area_in=np.full_like (ActualDuration_in, 1)*Brain_area/400
                Brain_size_in=np.full_like (ActualDuration_in, 1)*Brain_size/60
                total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr
                y_train = MaxAvgT[sub_list]
                X_traina = AvgPower_in
                #np.concatenate((AvgPower_in[:,:, np.newaxis], Thickness_in[:,:, np.newaxis]),axis=2)
                X_trainb=Thickness_in
                # X_trainc=SDR_in
                # X_traind = np.concatenate((Angle_in[:, :, np.newaxis], ElementPosition_in),axis=2)

                X_trainc=age_in[:, np.newaxis]
                X_traind=dis_in[:, np.newaxis]
                X_traine = ActualDuration_in[:, np.newaxis]
                X_trainf=ActualEnergy_in[:, np.newaxis]

                X_traing = Brain_area_in[:, np.newaxis]
                X_trainh=Brain_size_in[:, np.newaxis]
                X_traini = total_sdr_in[:, np.newaxis]

                y_train = y_train[:, np.newaxis]

                # 前向传播
                # 将训练数据转换为 PyTorch 的 Tensor 类型
                X_traina = torch.from_numpy(X_traina).float()
                X_trainb = torch.from_numpy(X_trainb).float()
                X_trainc = torch.from_numpy(X_trainc).float()
                X_traind = torch.from_numpy(X_traind).float()
                X_traine = torch.from_numpy(X_traine).float()
                X_trainf = torch.from_numpy(X_trainf).float()
                X_traing = torch.from_numpy(X_traing).float()
                X_trainh = torch.from_numpy(X_trainh).float()
                X_traini = torch.from_numpy(X_traini).float()
                y_train = torch.from_numpy(y_train).float()
                outputs = model(X_traina,X_trainb,X_trainc,X_traind,X_traine,X_trainf,X_traing,X_trainh,X_traini)

                loss = criterion(outputs, y_train)
                # loss = criterion(outputs[:, np.newaxis], y_train).detach_().requires_grad_(True)
                total_loss.append(loss)
            mean_loss = np.mean(total_loss)
        return mean_loss

    def train(train_list, test_list):
        CT_value_file = "./data/CT_value_20231219"
        model.train()
        CT_value = {}
        for file in train_list:
            ID = file.split('.')[0]
            k = 0
            path = os.path.join(CT_value_file, file)
            with open(path, 'r', encoding='UTF-8') as f:
                sub_data = []
                for line in f.readlines():
                    k = k + 1
                    if k < 4:
                        continue
                    line = line.strip('\n')  # 去掉换行符\n
                    data_list = line.split('\t')[1:]
                    data_list = list(map(float, data_list))  # 将每一行以空格为分隔符转换成列表
                    mea = sum(data_list) / len(data_list)
                    sub_data.append(mea)  # 全部除以1000
                sub_data = np.array(sub_data)
                CT_value[ID] = sub_data

        # SDR score of patient;patients x 1
        path = "./data/CT_proj_data20231216.xlsx"
        look_up_table_row_start = 2
        look_up_table_row_number = 90
        skull_area={}
        skull_score={}
        measurement={}
        ages={}
        disea={}
        look_up_table_excel = load_workbook(path)
        look_up_table_all_sheet = look_up_table_excel.sheetnames
        look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
        for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
            ID = look_up_table_sheet.cell(i, 1).value
            sub_data1 = look_up_table_sheet.cell(i, 8).value
            sub_data2 = look_up_table_sheet.cell(i, 11).value
            sub_data3 = look_up_table_sheet.cell(i, 17).value
            sub_data4 = look_up_table_sheet.cell(i, 20).value
            sub_data5 = look_up_table_sheet.cell(i, 22).value
            ages[ID]=sub_data1
            disea[ID]=sub_data2
            skull_area[ID] = sub_data3
            skull_score[ID] = sub_data4
            measurement[ID] = sub_data5

        Data_ver = "./data/Data_ver5"
        num_epochs = 10000
        best_val_loss = 1000  # initial
        for epoch in range(num_epochs):
            random.shuffle(train_list)
            # read .mat
            for file in train_list:
                CT_density = CT_value[file]
                Brain_area=skull_area[file]
                age=ages[file]
                dis=disea[file]
                total_sdr=skull_score[file]
                Brain_size=measurement[file]
                path = os.path.join(Data_ver, file)
                DATA_Mat = scio.loadmat(path)
                # print(DATA_Mat.keys())
                AvgPower = DATA_Mat['Elements']['AvgPower'][0][0][0, :]
                Thickness = DATA_Mat['Elements']['Thickness'][0][0][0, :]
                SDR = DATA_Mat['Elements']['SDR'][0][0][0, :]
                Angle = DATA_Mat['Elements']['Angle'][0][0][0, :]
                ElementPosition = DATA_Mat['Elements']['ElementPosition'][0][0][0, :]
                OnOff = DATA_Mat['Elements']['OnOff'][0][0][0, :]

                # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                # ActualEnergy = DATA_Mat['Sonication']['ActualEnergy'][0][0][:, 0]

                ActualDuration = DATA_Mat['Sonication']['PlannedDuration'][0][0][:, 0]#Sonication.PlannedDuration
                ActualEnergy = DATA_Mat['Sonication']['PlannedEnergy'][0][0][:, 0]
                MaxAvgT = DATA_Mat['Sonication']['MaxAvgT'][0][0][:, 0]  # max average temperature which our target.
                # num = random.randint(2 * int(len(AvgPower) / 3), len(AvgPower))
                # sub_list = random.sample(range(len(AvgPower)), num)
                # # sub_list = list(range(len(AvgPower)))
                # random.shuffle(sub_list)
                treatment_file = "./data/treatment_para/"
                path = os.path.join(treatment_file, file) + ".csv"
                data = pd.read_csv(path)
                aa = data.iloc[:, 19]
                # print(aa[1])
                indices_of_no = aa[aa == 'No             '].index
                indices_of_no=indices_of_no.tolist()
                num = random.randint(2 * int(len(indices_of_no) / 3), len(indices_of_no))
                sub_list = random.sample(indices_of_no, num)
                # random.sample(indices_of_no, num)


                AvgPower_in = []
                Thickness_in = []
                SDR_in = []
                Angle_in=[]
                ElementPosition_in=[]
                # ActualEnergy_in=[]

                MaxAvgT_out = []
                for m in sub_list:
                    AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                    Thickness_in.append(Thickness[m][:, 0] * CT_density * OnOff[m][:, 0])
                    SDR_in.append(SDR[m][:, 0]* OnOff[m][:, 0])  # 是否受能量启动影像
                    Angle_in.append(Angle[m][:, 0]*OnOff[m][:,0])
                    ElementPosition_in.append(ElementPosition[m][:, :] * OnOff[m][:, 0][:, np.newaxis].
                                              repeat(ElementPosition[m][:, :].shape[1], axis=1))
                    # AvgPower_in=AvgPower[m][:,0]
                    # Thickness_in=Thickness[m][:,0]
                    # SDR_in=SDR[m][:,0]
                    # OnOff_in=OnOff[m][:,0]

                    # ActualDuration_in.append(ActualDuration[m][:, 0])
                    # ActualEnergy_in.append(ActualEnergy[m][:,0])
                    # MaxAvgT_out.append(MaxAvgT[m][:,0])
                AvgPower_in = np.array(AvgPower_in)
                Thickness_in = np.array(Thickness_in)/1000
                SDR_in=np.array(SDR_in)
                SDR_in=-np.nan_to_num(SDR_in)
                Angle_in=np.array(Angle_in)/30
                ElementPosition_in = np.array(ElementPosition_in)/200

                ActualDuration_in = ActualDuration[sub_list]/10
                ActualEnergy_in = ActualEnergy[sub_list]/10000
                age_in = np.full_like (ActualDuration_in, 1)*age/100
                dis_in = np.full_like (ActualDuration_in, 1)*dis/40
                Brain_area_in=np.full_like (ActualDuration_in, 1)*Brain_area/400
                Brain_size_in=np.full_like (ActualDuration_in, 1)*Brain_size/60
                total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr
                y_train = MaxAvgT[sub_list]
                if random.random()>0.5:
                    x,y=AvgPower_in.shape
                    AvgPower_in=AvgPower_in+np.random.rand(x,y)/10
                    Thickness_in=Thickness_in+np.random.rand(x,y)
                    SDR_in=SDR_in+np.random.rand(x,y)/10
                    Angle_in = Angle_in + np.random.rand(x, y) / 10
                    ElementPosition_in = ElementPosition_in + np.random.rand(x, y,3) /10

                    age_in = age_in + np.random.rand(x) / 10
                    dis_in = dis_in + np.random.rand(x) / 10

                    ActualDuration_in=ActualDuration_in+np.random.rand(x)/10
                    ActualEnergy_in=ActualEnergy_in+np.random.rand(x)/10

                    Brain_area_in=Brain_area_in+np.random.rand(x)/10
                    Brain_size_in=Brain_size_in+np.random.rand(x)/10
                    total_sdr_in=total_sdr_in++np.random.rand(x)/10

                    y_train=y_train+np.random.rand(x)
                X_traina = AvgPower_in
                #np.concatenate((AvgPower_in[:,:, np.newaxis], Thickness_in[:,:, np.newaxis]),axis=2)
                X_trainb=Thickness_in
                # X_trainc=SDR_in
                # X_traind = np.concatenate((Angle_in[:, :, np.newaxis], ElementPosition_in),axis=2)
                X_trainc=age_in[:, np.newaxis]
                X_traind=dis_in[:, np.newaxis]

                X_traine = ActualDuration_in[:, np.newaxis]
                X_trainf=ActualEnergy_in[:, np.newaxis]
                X_traing = Brain_area_in[:, np.newaxis]
                X_trainh=Brain_size_in[:, np.newaxis]
                X_traini = total_sdr_in[:, np.newaxis]

                y_train = y_train[:, np.newaxis]

                # 前向传播
                # 将训练数据转换为 PyTorch 的 Tensor 类型
                X_traina = torch.from_numpy(X_traina).float()
                X_trainb = torch.from_numpy(X_trainb).float()
                X_trainc = torch.from_numpy(X_trainc).float()
                X_traind = torch.from_numpy(X_traind).float()
                X_traine = torch.from_numpy(X_traine).float()
                X_trainf = torch.from_numpy(X_trainf).float()
                X_traing = torch.from_numpy(X_traing).float()
                X_trainh = torch.from_numpy(X_trainh).float()
                X_traini = torch.from_numpy(X_traini).float()
                y_train = torch.from_numpy(y_train).float()
                outputs = model(X_traina,X_trainb,X_trainc,X_traind,X_traine,X_trainf,X_traing,X_trainh,X_traini)
                # y_train=sigmoid(y_train)
                loss = criterion(outputs, y_train)#.detach_().requires_grad_(True)

                # 反向传播和优化
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()

            if (epoch + 1) % 50 == 0:
                print(f'Epoch [{epoch + 1}/{num_epochs}], Loss: {loss.item()}')
                torch.save(model, latest_model)
                val_loss = val_test(test_list)
                print(f'Epoch [{epoch + 1}/{num_epochs}], val_Loss: {val_loss.item()}')
                if best_val_loss > val_loss:
                    best_val_loss = val_loss
                    print('Yayy! New best model saved!\n')
                    torch.save(model, best_model)
                print('\n')
    train(train_list,test_list)





