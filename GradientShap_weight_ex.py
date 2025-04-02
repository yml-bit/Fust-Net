import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import os
import scipy.io as scio
import random
import pandas as pd
from openpyxl import load_workbook
from captum.attr import IntegratedGradients
from captum.attr import DeepLift, GradientShap
import os
os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE"
import h5py
import torch.nn as nn
import torch.nn.functional as F

class MLPModel1(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel1, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(64, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))

        x = self.relu(self.fc1(x_E1))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel2(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel2, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)

        self.fc1 = nn.Linear(128, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, data):
        E=data[:,0].unsqueeze(dim=1)
        F = data[:, 1].unsqueeze(dim=1)
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))


        x_all = torch.cat((x_E1,x_F1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel3(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel3, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(192, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, data):
        E=data[:,0].unsqueeze(dim=1)
        F = data[:, 1].unsqueeze(dim=1)
        I = data[:, 2].unsqueeze(dim=1)
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(I))


        x_all = torch.cat((x_E1,x_F1,x_G1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel4(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel4, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(256, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, data):
        E=data[:,0].unsqueeze(dim=1)
        F = data[:, 1].unsqueeze(dim=1)
        H = data[:, 2].unsqueeze(dim=1)
        I = data[:, 3].unsqueeze(dim=1)
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(H))
        x_H1 = self.relu(self.fc_H1(I))

        x_all = torch.cat((x_E1,x_F1,x_G1,x_H1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel5(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel5, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(320, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, data):
        E=data[:,0].unsqueeze(dim=1)
        F = data[:,1].unsqueeze(dim=1)
        G = data[:, 2].unsqueeze(dim=1)
        H = data[:, 3].unsqueeze(dim=1)
        I = data[:, 4].unsqueeze(dim=1)
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel6(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel6, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)

        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, data):
        D=data[:,0].unsqueeze(dim=1)
        E = data[:, 1].unsqueeze(dim=1)
        F = data[:, 2].unsqueeze(dim=1)
        G = data[:, 3].unsqueeze(dim=1)
        H = data[:, 4].unsqueeze(dim=1)
        I = data[:, 5].unsqueeze(dim=1)
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        # x_C1 = self.relu(self.fc_C1(E))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_D1,x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

# A=AvgPower_in,B=Thickness_in,C=age_in,D=dis_in,E=ActualDuration_in,
# F=ActualEnergy_in,G=Brain_area_in,H=Brain_size_in,I=total_sdr_in
class AttentionLayer(nn.Module):
    def __init__(self, hidden_dim):
        super(AttentionLayer, self).__init__()
        self.W = nn.Linear(hidden_dim,hidden_dim)
        self.V = nn.Linear(hidden_dim, 1)
        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        # weighted_x = x+x * attn_weights  # Weighted sum of features

        attn_weights = self.dropout(attn_weights)
        weighted_x = self.layer_norm(x + x * attn_weights)
        return weighted_x  # Sum across features

class FactorAttentionLayer(nn.Module):
    def __init__(self, hidden_dim):
        super(FactorAttentionLayer, self).__init__()
        self.W = nn.Linear(hidden_dim, hidden_dim)
        self.V = nn.Linear(hidden_dim, 1)
        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        # weighted_x = x + x * attn_weights  # Weighted sum of features

        attn_weights = self.dropout(attn_weights)
        weighted_x = self.layer_norm(x + x * attn_weights)

        return weighted_x  # Sum across features

class MultiHeadFactorAttentionLayer(nn.Module):
    def __init__(self, hidden_dim, num_heads=7):
        super(MultiHeadFactorAttentionLayer, self).__init__()
        assert hidden_dim % num_heads == 0, "Hidden dimension must be divisible by number of heads."

        self.hidden_dim = hidden_dim
        self.num_heads = num_heads
        self.head_dim = hidden_dim // num_heads

        self.W = nn.Linear(hidden_dim, hidden_dim)
        self.V = nn.Linear(hidden_dim, num_heads)

        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        batch_size, num_features, _ = x.size()

        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        attn_weights = attn_weights.view(batch_size, num_features, self.num_heads, 1)

        # Split the last dimension into (num_heads, head_dim)
        u_split = u.view(batch_size, num_features, self.num_heads, self.head_dim)

        # Weighted sum of features
        weighted_x = (u_split * attn_weights).sum(dim=2)

        # Residual connection and layer normalization
        weighted_x = self.dropout(weighted_x)
        weighted_x = self.layer_norm(x + x*weighted_x)#5.8
        # weighted_x=x + x * weighted_x # 6.021325588226318

        return weighted_x

class MLPModel7(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel7, self).__init__()

        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        # self.factor_attention2 = FactorAttentionLayer(input_dim)
        # self.factor_attention = MultiHeadFactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, array):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        # 通过对应的线性层处理每个输入特征
        n = array.shape[0]
        inputs = [col.reshape((n, 1)) for col in array.T]
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs)]
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
        # x_all=x_all.reshape(x_all.shape[0], -1)

        # 应用注意力机制
        xx = self.factor_attention(x_all.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#attention2；+dropout layernorm+5.455
        # xx = self.attention(x_all).reshape(x_all.shape[0], -1)#attention3

        # xx = self.factor_attention(x_all.permute(0, 2, 1))
        # xx = self.attention(xx.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#mutil-attention

        # 应用残差连接
        x = self.fc_hidden(xx)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

if __name__ == '__main__':
    cross = 5
    models = ["model1_dur", "model2_dur_energy", "model3_dur_energy_sdr", "model4_dur_energy_size_sdr",
              "model5_dur_energy_area_size_sdr","model6_ang_dur_energy_area_size_sdr", "model6_ang_dur_energy_area_size_sdr2",
              "model7_dis_dur_energy_area_size_sdr_avp_seq"]
    for j in range(len(models)):
        if j<7:
            continue
        mm = models[j]
        # f_weight = open(weight_record, "a",encoding='utf-8')
        script_dir = os.path.dirname(os.path.abspath(__file__))
        weight_record = os.path.join(script_dir, 'GradientShape', "GradientShap_weight" + str(j + 1) + ".txt")
        directory = os.path.dirname(weight_record)
        if not os.path.exists(directory):
            os.makedirs(directory)
        # with open(weight_record, "a", encoding='utf-8') as f_weight:
        #     pass
        f_weight = open(weight_record, "a", encoding='utf-8')
        attributions_list = []
        for i in range(cross):
            # if i < 4:
            #     continue
            fold = "./fold_" + str(i + 1)
            os.makedirs(fold, exist_ok=True)
            model_path = fold + "/" + mm
            # latest_model = model_path + '/latest_model.pth'
            best_model = model_path + '/best_model.pth'
            model = torch.load(best_model)
            # with open(f"./flist/test_fold_{i + 1}.txt", "r") as f_test:
            #     f_test_list = f_test.read().split("\n")

            with open("external_list.txt", "r") as f_test:
                f_test_list = f_test.read().split("\n")

            test_list = []
            for file in f_test_list:
                if file:  # 检查是否为空字符串
                    test_list.append(file.split("\n")[0])

            ig = GradientShap(model)
            # criterion = nn.MSELoss()
            criterion = nn.L1Loss(reduction="none")

            model.eval()
            with torch.no_grad():
                CT_value_file = "./data/DrXiong_files_3"
                CT_value = {}
                attributions_list = []
                for file in test_list:
                    ID = file.split('.')[0].split('_')[1]

                path = "./data/CT_proj_data20240909.xlsx"
                look_up_table_row_start = 2
                look_up_table_row_number = 90
                skull_area = {}
                skull_score = {}
                measurement = {}
                ages = {}
                disea = {}
                look_up_table_excel = load_workbook(path)
                look_up_table_all_sheet = look_up_table_excel.sheetnames
                look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
                for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
                    ID = str(look_up_table_sheet.cell(i, 1).value)
                    sub_data1 = look_up_table_sheet.cell(i, 7).value
                    sub_data2 = look_up_table_sheet.cell(i, 10).value
                    sub_data3 = look_up_table_sheet.cell(i, 15).value
                    sub_data4 = look_up_table_sheet.cell(i, 17).value
                    sub_data5 = look_up_table_sheet.cell(i, 19).value
                    ages[ID] = sub_data1
                    disea[ID] = sub_data2
                    skull_area[ID] = sub_data3
                    skull_score[ID] = sub_data4
                    measurement[ID] = sub_data5

            total_loss = []
            Data_ver = "./data/DrXiong_files_3"
            for files in test_list:
                file = str(files.split('.')[0].split('_')[1])
                # CT_density = CT_value[file]
                Brain_area = skull_area[file]
                age = ages[file]
                dis = disea[file]
                total_sdr = skull_score[file]
                Brain_size = measurement[file]
                path = os.path.join(Data_ver, files)
                DATA_Mat = scio.loadmat(path)
                # print(DATA_Mat.keys())
                st = "Tx_" + file
                AvgPower = DATA_Mat[st]['AvgPower'][0][0][0, :]
                Thickness = DATA_Mat[st]['Thickness'][0][0][0, :]
                SDR = DATA_Mat[st]['SDR'][0, 0]
                Angle = DATA_Mat[st]['Angle'][0][0][0, :]
                # ElementPosition = DATA_Mat[st]['ElementPosition'][0][0][:,0]
                OnOff = DATA_Mat[st]['OnEl'][0][0][0, :]

                # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                ActualEnergy = DATA_Mat[st]['PlannedEnergy'][0][0][:, 0]
                ActualDuration = DATA_Mat[st]['PlannedDuration'][0][0][:, 0]  # Sonication.PlannedDuration
                MaxAvgT = DATA_Mat[st]['TempAvg'][0][0][:, 0]  # max average temperature which our target.
                # sub_list = list(range(len(AvgPower)))

                treatment_file = "./data/TreatSummary/"
                path = os.path.join(treatment_file, file) + "/TreatSummary.csv"
                data = pd.read_csv(path)
                aa = data.iloc[:, 19]
                # print(aa[1])
                sub_list = aa[aa == 'No             '].index

                AvgPower_in = []
                Thickness_in = []
                SDR_in = []
                Angle_in = []
                ElementPosition_in = []
                # ActualEnergy_in=[]

                MaxAvgT_out = []
                for m in sub_list:
                    AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                    # Thickness_in.append(Thickness[m][:, 0] * CT_density * OnOff[m][:, 0])
                    # SDR_in.append(SDR[m][:, 0] * OnOff[m][:, 0])  # 是否受能量启动影像
                    # Angle_in.append(Angle[m][:, 0] * OnOff[m][:, 0])
                    # ElementPosition_in.append(ElementPosition[m][:, :] * OnOff[m][:, 0][:, np.newaxis].
                    #                           repeat(ElementPosition[m][:, :].shape[1], axis=1))
                AvgPower_in = np.array(AvgPower_in)
                Thickness_in = np.array(Thickness_in) / 20000
                # SDR_in = np.array(SDR_in)
                # SDR_in = -np.nan_to_num(SDR_in)
                Angle_in = np.array(Angle_in) / 30
                ElementPosition_in = np.array(ElementPosition_in) / 200

                ActualDuration_in = ActualDuration[sub_list] / 10
                ActualEnergy_in = ActualEnergy[sub_list] / 10000
                age_in = np.full_like(ActualDuration_in, 1) * age / 100
                dis_in = np.full_like(ActualDuration_in, 1) * dis / 40

                Brain_area_in = np.full_like(ActualDuration_in, 1) * Brain_area / 400
                Brain_size_in = np.full_like(ActualDuration_in, 1) * Brain_size / 60
                total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr
                y_train0 = MaxAvgT[sub_list]
                X_traina = AvgPower_in
                X_traina = np.mean(X_traina, axis=1).reshape(-1, 1)
                # X_trainb = Thickness_in
                # X_trainb = np.mean(X_trainb, axis=1).reshape(-1, 1)

                X_trainc = age_in[:, np.newaxis]
                X_traind = dis_in[:, np.newaxis]
                X_traine = ActualDuration_in[:, np.newaxis]
                X_trainf = ActualEnergy_in[:, np.newaxis]

                X_traing = Brain_area_in[:, np.newaxis]
                X_trainh = Brain_size_in[:, np.newaxis]
                X_traini = total_sdr_in[:, np.newaxis]
                x_trainz = np.array(sub_list)[:, np.newaxis] / 20

                y_train0 = y_train0[:, np.newaxis]

                # 前向传播
                # 将训练数据转换为 PyTorch 的 Tensor 类型
                X_traina = torch.from_numpy(X_traina).float()
                # X_trainb = torch.from_numpy(X_trainb).float()
                X_trainc = torch.from_numpy(X_trainc).float()
                X_traind = torch.from_numpy(X_traind).float()
                X_traine = torch.from_numpy(X_traine).float()
                X_trainf = torch.from_numpy(X_trainf).float()
                X_traing = torch.from_numpy(X_traing).float()
                X_trainh = torch.from_numpy(X_trainh).float()
                X_traini = torch.from_numpy(X_traini).float()
                x_trainz = torch.from_numpy(x_trainz).float()

                y_train = torch.from_numpy(y_train0).float()
                # outputs = model(X_traina,X_trainb,X_trainc,X_traind,X_traine,X_trainf,X_traing,X_trainh,X_traini)
                if j==2:
                    input = torch.cat((X_traine, X_trainf), dim=1)
                elif j==3:
                    input = torch.cat((X_traine, X_trainf, X_traini), dim=1)
                elif j == 4:
                    input = torch.cat((X_traine, X_trainf, X_trainh,X_traini), dim=1)
                elif j==5:
                    input = torch.cat((X_traine, X_trainf, X_traing, X_trainh, X_traini), dim=1)
                elif j==6:
                    input=torch.cat((X_traind,X_traine,X_trainf,X_traing,X_trainh,X_traini), dim=1)
                elif j==7:
                    input = torch.cat(
                        (X_traine, X_trainf, X_traini, X_traina, X_trainh, X_traing, X_traind, x_trainz), dim=1)

                outputs = model(input)
                tensor_output = model(input)
                attributions = ig.attribute(input, baselines=torch.zeros_like(input))
                weight = attributions.mean(dim=0).numpy()[np.newaxis, :]
                attributions_list.append(weight)
                f_weight.write(str(weight) + '\n')
    mean_weight = np.mean(np.array(attributions_list),axis=0)
    f_weight.write(str(mean_weight) + '\n')
    # f_weight.close()#model5_dur_energy_area_size_sdr