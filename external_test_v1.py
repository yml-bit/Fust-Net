import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import os
import scipy.io as scio
import random
import pandas as pd
from openpyxl import load_workbook
import h5py
from scipy import stats

class MLPModel1(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel1, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(64, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))

        x = self.relu(self.fc1(x_E1))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel2(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel2, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)

        self.fc1 = nn.Linear(128, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))


        x_all = torch.cat((x_E1,x_F1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel3(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel3, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(192, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(I))


        x_all = torch.cat((x_E1,x_F1,x_G1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel4(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel4, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(256, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(H))
        x_H1 = self.relu(self.fc_H1(I))


        x_all = torch.cat((x_E1,x_F1,x_G1,x_H1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel5(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel5, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(320, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel6(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel6, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1024, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        # x_C1 = self.relu(self.fc_C1(E))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_D1,x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

# A=AvgPower_in,B=Thickness_in,C=age_in,D=dis_in,E=ActualDuration_in,
# F=ActualEnergy_in,G=Brain_area_in,H=Brain_size_in,I=total_sdr_in
if __name__ == '__main__':
    # random split
    cross = 5
    models = ["model1_dur", "model2_dur_energy", "model3_dur_energy_sdr", "model4_dur_energy_size_sdr",
              "model5_dur_energy_area_size_sdr","model6_ang_dur_energy_area_size_sdr", "model6_ang_dur_energy_area_size_sdr2"]
    for j in range(len(models)):
        mm = models[j]
        if j<6:
            continue
        for i in range(cross):
            # if i < 4:
            #     continue
            fold = "./fold_" + str(i + 1)
            os.makedirs(fold, exist_ok=True)
            model_path = fold + "/" + mm
            # latest_model = model_path + '/latest_model.pth'
            best_model = model_path + '/best_model.pth'
            in_results = fold + '/external_results/'
            if not os.path.exists(in_results):
                os.mkdir(in_results)
            test_record = in_results + mm + ".txt"
            loss_record = in_results + mm + "_loss.txt"
            # ex_results = fold + '/external_results/'
            if os.path.exists(test_record):
                os.remove(test_record)
                os.remove(loss_record)
            else:
                pass  # 文件不存在，跳过删除操作
            # 读取数据
            with open("external_list.txt", "r") as f_test:
                f_test_list = f_test.read().split("\n")

            test_list = []
            for file in f_test_list:
                if file:  # 检查是否为空字符串
                    test_list.append(file.split("\n")[0])

            model = torch.load(best_model)  #
            criterion = nn.L1Loss(reduction="none")

            model.eval()
            with torch.no_grad():
                f_test = open(test_record, "w")
                loss_test = open(loss_record, "w")
                CT_value_file = "./data/DrXiong_files_3"
                CT_value = {}
                for file in test_list:
                    ID = file.split('.')[0].split('_')[1]
                    # SDR score of patient;patients x 1
                    path = "./data/CT_proj_data20240909.xlsx"
                    look_up_table_row_start = 2
                    look_up_table_row_number = 90
                    skull_area = {}
                    skull_score = {}
                    measurement = {}
                    ages = {}
                    disea = {}
                    look_up_table_excel = load_workbook(path)
                    look_up_table_all_sheet = look_up_table_excel.sheetnames
                    look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
                    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
                        ID = str(look_up_table_sheet.cell(i, 1).value)
                        sub_data1 = look_up_table_sheet.cell(i, 7).value
                        sub_data2 = look_up_table_sheet.cell(i, 10).value
                        sub_data3 = look_up_table_sheet.cell(i, 15).value
                        sub_data4 = look_up_table_sheet.cell(i, 17).value
                        sub_data5 = look_up_table_sheet.cell(i, 19).value
                        ages[ID] = sub_data1
                        disea[ID] = sub_data2
                        skull_area[ID] = sub_data3
                        skull_score[ID] = sub_data4
                        measurement[ID] = sub_data5

                total_loss = []
                Data_ver = "./data/DrXiong_files_3"
                for files in test_list:
                    file = str(files.split('.')[0].split('_')[1])
                    # CT_density = CT_value[file]
                    Brain_area = skull_area[file]
                    age = ages[file]
                    dis = disea[file]
                    total_sdr = skull_score[file]
                    Brain_size = measurement[file]
                    path = os.path.join(Data_ver, files)
                    DATA_Mat = scio.loadmat(path)
                    # print(DATA_Mat.keys())
                    st="Tx_"+file
                    AvgPower = DATA_Mat[st]['AvgPower'][0][0][0,:]
                    Thickness = DATA_Mat[st]['Thickness'][0][0][0,:]
                    SDR = DATA_Mat[st]['SDR'][0,0]
                    Angle = DATA_Mat[st]['Angle'][0][0][0,:]
                    # ElementPosition = DATA_Mat[st]['ElementPosition'][0][0][:,0]
                    OnOff = DATA_Mat[st]['OnEl'][0][0][0,:]

                    # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                    ActualEnergy = DATA_Mat[st]['PlannedEnergy'][0][0][:, 0]
                    ActualDuration = DATA_Mat[st]['PlannedDuration'][0][0][:, 0] # Sonication.PlannedDuration
                    MaxAvgT = DATA_Mat[st]['TempAvg'][0][0][:, 0]  # max average temperature which our target.
                    # sub_list = list(range(len(AvgPower)))

                    treatment_file = "./data/TreatSummary/"
                    path = os.path.join(treatment_file, file) + "/TreatSummary.csv"
                    data = pd.read_csv(path)
                    aa = data.iloc[:, 19]
                    # print(aa[1])
                    sub_list = aa[aa == 'No             '].index

                    AvgPower_in = []
                    Thickness_in = []
                    SDR_in = []
                    Angle_in = []
                    ElementPosition_in = []
                    # ActualEnergy_in=[]

                    MaxAvgT_out = []
                    for m in sub_list:
                        AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                        # SDR_in.append(SDR[m][:, 0] * OnOff[m][:, 0])  # 是否受能量启动影像
                        Angle_in.append(Angle[m][:, 0] * OnOff[m][:, 0])

                    AvgPower_in = np.array(AvgPower_in)
                    # SDR_in = np.array(SDR_in)
                    # SDR_in = -np.nan_to_num(SDR_in)
                    Angle_in = np.array(Angle_in) / 30

                    ActualDuration_in = ActualDuration[sub_list] / 10
                    ActualEnergy_in = ActualEnergy[sub_list] / 10000
                    age_in = np.full_like(ActualDuration_in, 1) * age / 100
                    dis_in = np.full_like(ActualDuration_in, 1) * dis / 40

                    Brain_area_in = np.full_like(ActualDuration_in, 1) * Brain_area / 400
                    Brain_size_in = np.full_like(ActualDuration_in, 1) * Brain_size / 60
                    total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr

                    y_train0 = MaxAvgT[sub_list]
                    X_traina = AvgPower_in
                    X_trainb = X_traina
                    X_trainc = age_in[:, np.newaxis]
                    # X_traind = Angle_in#X_traina
                    X_traind = dis_in[:, np.newaxis]
                    X_traine = ActualDuration_in[:, np.newaxis]
                    X_trainf = ActualEnergy_in[:, np.newaxis]
                    X_traing = Brain_area_in[:, np.newaxis]
                    X_trainh = Brain_size_in[:, np.newaxis]
                    X_traini = total_sdr_in[:, np.newaxis]
                    #A=AvgPower_in,B=Thickness_in,C=age_in,D=dis_in,E=ActualDuration_in,
                    # F=ActualEnergy_in,G=Brain_area_in,H=Brain_size_in,I=total_sdr_in
                    y_train0 = y_train0[:, np.newaxis]
                    # 前向传播
                    # 将训练数据转换为 PyTorch 的 Tensor 类型
                    X_traina = torch.from_numpy(X_traina).float()
                    X_trainb = torch.from_numpy(X_trainb).float()
                    X_trainc = torch.from_numpy(X_trainc).float()
                    X_traind = torch.from_numpy(X_traind).float()
                    X_traine = torch.from_numpy(X_traine).float()
                    X_trainf = torch.from_numpy(X_trainf).float()
                    X_traing = torch.from_numpy(X_traing).float()
                    X_trainh = torch.from_numpy(X_trainh).float()
                    X_traini = torch.from_numpy(X_traini).float()

                    y_train = torch.from_numpy(y_train0).float()
                    outputs = model(X_traina,X_trainb,X_trainc,X_traind,X_traine,X_trainf,X_traing,X_trainh,X_traini)
                    absolute_errors = criterion(outputs, y_train)
                    absolute_errors = np.array(absolute_errors)[:, 0]
                    # absolute_errors = torch.abs(outputs[:,0], y_train[:,0])
                    mean_absolute_error = absolute_errors.mean()
                    total_loss.append(mean_absolute_error)
                    sample_variance = absolute_errors.var()
                    loss = str(mean_absolute_error)  #
                    predict = outputs.detach().numpy()
                    out = np.hstack((y_train0, predict))  # target+predict
                    # f_test.writelines(out)
                    for ii in range(len(out)):
                        f_test.write(str(out[ii]) + '\n')
                    f_test.writelines("\n")
                    loss_test.write(str(loss) + '\n')
                errors = np.array([total_loss])
                mean_error = np.mean(errors)  # 计算均值和标准差
                std_error = np.std(errors, ddof=1)
                n = len(errors)  # 样本量
                se = std_error / np.sqrt(n)  # 计算标准误
                # 确定z值（95%置信水平）
                z_value = stats.norm.ppf(0.975)  # 对应95%置信水平的z值
                lower_bound = mean_error - z_value * se  # 计算置信区间
                upper_bound = mean_error + z_value * se
                # 将均值和置信区间追加写入 f_test 文件
                f_test.write(f"Mean Error: {mean_error:.4f}\n")
                f_test.write(f"95% Confidence Interval: [{lower_bound:.4f}, {upper_bound:.4f}]\n")
                f_test.write("\n")  # 添加一个空行分隔不同的结果块