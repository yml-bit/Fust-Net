import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import os
import scipy.io as scio
import random
import pandas as pd
from openpyxl import load_workbook
import h5py
import scipy
import torch
import torch.nn as nn
import torch.nn.functional as F

def set_seed(seed):
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def total_external_data():
    with open("external_list.txt", "r") as tf:
        test_list = tf.read().split("\n")
    treatment_file = "./data/TreatSummary/"
    total1=0
    total2=0
    sub_data = {}
    for files in test_list:
        file = str(files.split('.')[0].split('_')[1])
        path = os.path.join(treatment_file, file) + "/TreatSummary.csv"
        data = pd.read_csv(path)
        sub_data["Flag"] = data.iloc[:, 19]
        aa=data.iloc[:, 19]
        total1 += 1
        total2+=len(aa)
    print("patients number:%.2f"%total1)
    print("treatment number:%.2f"%total2)

def actual_external_data():
    with open("external_list.txt", "r") as tf:
        test_list = tf.read().split("\n")
    treatment_file = "./data/TreatSummary/"
    patient=0
    total=0
    sub_data = {}
    for files in test_list:
        file = str(files.split('.')[0].split('_')[1])
        path = os.path.join(treatment_file, file) + "/TreatSummary.csv"
        data = pd.read_csv(path)
        sub_data["Flag"] = data.iloc[:, 19]
        aa=data.iloc[:, 19]
        # print(aa[1])
        indices_of_no = aa[aa =='No             '].index
        if len(indices_of_no)==len(aa):
            patient+=1
        else:
            a=1
        total+=len(indices_of_no)
    print("effective patients:%.2f"%patient)
    print("effective number:%.2f"%total)

class MLPModel6(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel6, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        # x_C1 = self.relu(self.fc_C1(E))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_D1,x_E1,x_F1,x_G1,x_H1,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class MLPModel66(nn.Module):
    def __init__(self, input_dim, output_dim):
        super(MLPModel66, self).__init__()

        # self.fc_A1 = nn.Linear(1024, 512)
        # self.fc_B1 = nn.Linear(1024, 512)
        # 辅助分支
        # self.fc_C1 = nn.Linear(1, 64)
        self.fc_D1 = nn.Linear(1, 64)
        self.fc_E1 = nn.Linear(1, 64)
        self.fc_F1 = nn.Linear(1, 64)
        self.fc_G1 = nn.Linear(1, 64)
        self.fc_H1 = nn.Linear(1, 64)
        self.fc_I1 = nn.Linear(1, 64)


        self.fc1 = nn.Linear(384, 64)
        self.fc2 = nn.Linear(64, output_dim)  # 隐藏层到输出层
        self.relu = nn.ReLU()  # 非线性激活函数

    def forward(self, A,B, C, D, E,F,G,H,I):
        # x_A1 = self.relu(self.fc_A1(A))
        # x_B1 = self.relu(self.fc_B1(C))
        # x_C1 = self.relu(self.fc_C1(E))
        x_D1 = self.relu(self.fc_D1(D))
        x_E1 = self.relu(self.fc_E1(E))
        x_F1 = self.relu(self.fc_F1(F))
        x_G1 = self.relu(self.fc_G1(G))
        x_H1 = self.relu(self.fc_H1(H))
        x_I1 = self.relu(self.fc_I1(I))

        x_all = torch.cat((x_D1,x_E1*0,x_F1,x_G1*0,x_H1*0,x_I1), dim=1)
        x = self.relu(self.fc1(x_all))
        # x = self.relu(self.fc2(x))
        x = self.fc2(x)
        return x

class AttentionLayer(nn.Module):
    def __init__(self, hidden_dim):
        super(AttentionLayer, self).__init__()
        self.W = nn.Linear(hidden_dim,hidden_dim)
        self.V = nn.Linear(hidden_dim, 1)
        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        # weighted_x = x+x * attn_weights  # Weighted sum of features

        attn_weights = self.dropout(attn_weights)
        weighted_x = self.layer_norm(x + x * attn_weights)
        return weighted_x  # Sum across features

# class FactorAttentionLayer(nn.Module):
#     def __init__(self, hidden_dim):
#         super(FactorAttentionLayer, self).__init__()
#         self.W = nn.Linear(hidden_dim, hidden_dim)
#         self.V = nn.Linear(hidden_dim, 1)
#
#     def forward(self, x):
#         # x shape: (batch_size, num_features, hidden_dim)
#         u = torch.tanh(self.W(x))  # Apply non-linearity
#         attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
#         weighted_x = x+x * attn_weights  # Weighted sum of features
#
#         return weighted_x  # Sum across features

class FactorAttentionLayer(nn.Module):
    def __init__(self, hidden_dim):
        super(FactorAttentionLayer, self).__init__()
        self.W = nn.Linear(hidden_dim, hidden_dim)
        self.V = nn.Linear(hidden_dim, 1)
        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        # weighted_x = x + x * attn_weights  # Weighted sum of features

        attn_weights = self.dropout(attn_weights)
        weighted_x = self.layer_norm(x + x * attn_weights)

        return weighted_x  # Sum across features

class MultiHeadFactorAttentionLayer(nn.Module):
    def __init__(self, hidden_dim, num_heads=7):
        super(MultiHeadFactorAttentionLayer, self).__init__()
        assert hidden_dim % num_heads == 0, "Hidden dimension must be divisible by number of heads."

        self.hidden_dim = hidden_dim
        self.num_heads = num_heads
        self.head_dim = hidden_dim // num_heads

        self.W = nn.Linear(hidden_dim, hidden_dim)
        self.V = nn.Linear(hidden_dim, num_heads)

        self.layer_norm = nn.LayerNorm(hidden_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        batch_size, num_features, _ = x.size()

        # x shape: (batch_size, num_features, hidden_dim)
        u = torch.tanh(self.W(x))  # Apply non-linearity
        attn_weights = F.softmax(self.V(u), dim=1)  # Compute attention weights
        attn_weights = attn_weights.view(batch_size, num_features, self.num_heads, 1)

        # Split the last dimension into (num_heads, head_dim)
        u_split = u.view(batch_size, num_features, self.num_heads, self.head_dim)

        # Weighted sum of features
        weighted_x = (u_split * attn_weights).sum(dim=2)

        # Residual connection and layer normalization
        weighted_x = self.dropout(weighted_x)
        weighted_x = self.layer_norm(x + x*weighted_x)#5.8
        # weighted_x=x + x * weighted_x # 6.021325588226318

        return weighted_x

#只用一个注意力
class MLPModel7(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel7, self).__init__()

        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        # self.factor_attention2 = FactorAttentionLayer(input_dim)
        # self.factor_attention = MultiHeadFactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, inputs):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        # 通过对应的线性层处理每个输入特征
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs)]
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
        # x_all=x_all.reshape(x_all.shape[0], -1)

        # 应用注意力机制
        xx = self.factor_attention(x_all.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#attention2；+dropout layernorm+5.455
        # xx = self.attention(x_all).reshape(x_all.shape[0], -1)#attention3

        # xx = self.factor_attention(x_all.permute(0, 2, 1))
        # xx = self.attention(xx.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#mutil-attention

        # 应用残差连接
        x = self.fc_hidden(xx)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

class MLPModel8(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel8, self).__init__()
        self.fc=nn.Linear(1024, hidden_dim)
        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        # self.factor_attention2 = FactorAttentionLayer(input_dim)
        # self.factor_attention = MultiHeadFactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, inputs):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        fc = self.relu(self.fc(inputs[0]))
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs[1:])]
        processed_inputs=[fc]+processed_inputs
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
        # x_all=x_all.reshape(x_all.shape[0], -1)

        # 应用注意力机制
        xx = self.factor_attention(x_all.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#attention2；+dropout layernorm+5.455
        # xx = self.attention(x_all).reshape(x_all.shape[0], -1)#attention3

        # xx = self.factor_attention(x_all.permute(0, 2, 1))
        # xx = self.attention(xx.permute(0, 2, 1)).reshape(x_all.shape[0], -1)#mutil-attention

        # 应用残差连接
        x = self.fc_hidden(xx)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

####################下面为调试模型，不是最优的模型###############
#无注意力机制
class MLPModel7a(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel7a, self).__init__()

        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, inputs):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        # 通过对应的线性层处理每个输入特征
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs)]

        # 拼接所有处理过的输入
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)
        x_all=x_all.reshape(x_all.shape[0], -1)
        # 应用注意力机制
        # attended_features = self.attention(x_all)
        # attended_factors = self.factor_attention(attended_features.permute(0, 2, 1)).reshape(x_all.shape[0], -1)

        # 应用残差连接
        x = self.fc_hidden(x_all)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

#用了两次注意力
class MLPModel7b(nn.Module):#m1
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(MLPModel7b, self).__init__()

        self.fc_input = nn.ModuleList([nn.Linear(1, hidden_dim) for _ in range(input_dim)])
        self.attention = AttentionLayer(hidden_dim)
        self.factor_attention = FactorAttentionLayer(input_dim)
        self.fc_hidden = nn.Linear(input_dim*hidden_dim, hidden_dim)
        self.fc_output = nn.Linear(hidden_dim, output_dim)
        self.relu = nn.ReLU()
        self.dropout = nn.Dropout(0.2)

    def forward(self, inputs):
        # 假设 `inputs` 是一个元组或列表，其中每个元素对应一个输入特征
        # 通过对应的线性层处理每个输入特征
        processed_inputs = [self.relu(fc(input)) for fc, input in zip(self.fc_input, inputs)]

        # 拼接所有处理过的输入
        x_all = torch.stack(processed_inputs, dim=1)  # (batch_size, num_features, hidden_dim)

        # 应用注意力机制
        attended_features = self.attention(x_all)
        attended_factors = self.factor_attention(attended_features.permute(0, 2, 1)).reshape(x_all.shape[0], -1)

        # 应用残差连接
        x = self.fc_hidden(attended_factors)
        x = self.relu(x)
        x = self.dropout(x)

        # 最终输出
        x = self.fc_output(x)

        return x

if __name__ == '__main__':
    set_seed(64)  # 可以选择任何整数值作为随机种子
    # total_external_data()#统计需要
    # actual_external_data()#统计需要
    cross = 5
    models = ["model7_dis_dur_energy_area_size_sdr_avp_seq","model8_ang_dis_dur_energy_area_size_sdr_avp_seq"]
    for j in range(len(models)):
        mm = models[j]
        if j != 0:
            continue
        for i in range(cross):
            if i>2:
                continue

            # if i<3:
            #     continue
            fold = "./fold_" + str(i + 1)
            os.makedirs(fold, exist_ok=True)
            model_path = fold + "/" + mm
            if not os.path.exists(model_path):
                os.mkdir(model_path)
            latest_model = model_path + '/latest_model.pth'
            best_model = model_path + '/best_model.pth'
            # 读取数据
            with open(f"./flist/train_fold_{i + 1}.txt", "r") as f_train, open(f"./flist/test_fold_{i + 1}.txt",
                                                                               "r") as f_test:
                f_train_list = f_train.read().split("\n")
                f_test_list = f_test.read().split("\n")
            train_list = []
            for file in f_train_list:
                if file:  # 检查是否为空字符串
                    train_list.append(file.split("\n")[0])

            test_list = []
            for file in f_test_list:
                if file:  # 检查是否为空字符串
                    test_list.append(file.split("\n")[0])

            input_dim = 8  # 8 9
            hidden_dim = 64  # 64
            output_dim = 1  # Predicted temperature
            # model = MLPModel7(input_dim, hidden_dim, output_dim)
            model = MLPModel8(9, hidden_dim, output_dim)
            optimizer = optim.Adam(model.parameters(), lr=5e-4, weight_decay=0.01)  # weight_decay即为L2正则化系数
            criterion=nn.MSELoss()

            def val_test(test_list):
                model.eval()
                with torch.no_grad():
                    # 使用训练好的模型进行预测
                    # 假设 test_data 是测试数据
                    test_record = "test_record.txt"
                    f_test = open(test_record, "w")

                    CT_value_file = "./data/CT_value_20231219"
                    CT_value = {}
                    for file in test_list:
                        ID = file.split('.')[0]
                        k = 0
                        path = os.path.join(CT_value_file, file)
                        with open(path, 'r', encoding='UTF-8') as f:
                            sub_data = []
                            for line in f.readlines():
                                k = k + 1
                                if k < 4:
                                    continue
                                line = line.strip('\n')  # 去掉换行符\n
                                data_list = line.split('\t')[1:]
                                data_list = list(map(float, data_list))  # 将每一行以空格为分隔符转换成列表
                                mea = sum(data_list) / len(data_list)
                                sub_data.append(mea)  # 全部除以1000
                            sub_data = np.array(sub_data)
                            CT_value[ID] = sub_data

                    # SDR score of patient;patients x 1
                    path = "./data/CT_proj_data20231216.xlsx"
                    look_up_table_row_start = 2
                    look_up_table_row_number = 90
                    skull_area = {}
                    skull_score = {}
                    measurement = {}
                    ages = {}
                    disea = {}
                    look_up_table_excel = load_workbook(path,read_only=True)
                    look_up_table_all_sheet = look_up_table_excel.sheetnames
                    look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
                    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
                        ID = look_up_table_sheet.cell(i, 1).value
                        sub_data1 = look_up_table_sheet.cell(i, 8).value
                        sub_data2 = look_up_table_sheet.cell(i, 11).value
                        sub_data3 = look_up_table_sheet.cell(i, 17).value
                        sub_data4 = look_up_table_sheet.cell(i, 20).value
                        sub_data5 = look_up_table_sheet.cell(i, 22).value
                        ages[ID] = sub_data1
                        disea[ID] = sub_data2
                        skull_area[ID] = sub_data3
                        skull_score[ID] = sub_data4
                        measurement[ID] = sub_data5

                    total_loss = []
                    Data_ver = "./data/Data_ver5"
                    for file in test_list:
                        CT_density = CT_value[file]
                        Brain_area=skull_area[file]
                        age=ages[file]
                        dis=disea[file]
                        total_sdr=skull_score[file]
                        Brain_size=measurement[file]
                        path = os.path.join(Data_ver, file)
                        DATA_Mat = scio.loadmat(path)
                        # print(DATA_Mat.keys())
                        AvgPower = DATA_Mat['Elements']['AvgPower'][0][0][0, :]
                        Thickness = DATA_Mat['Elements']['Thickness'][0][0][0, :]
                        SDR = DATA_Mat['Elements']['SDR'][0][0][0, :]
                        Angle = DATA_Mat['Elements']['Angle'][0][0][0, :]
                        ElementPosition = DATA_Mat['Elements']['ElementPosition'][0][0][0, :]
                        OnOff = DATA_Mat['Elements']['OnOff'][0][0][0, :]

                        # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                        # ActualEnergy = DATA_Mat['Sonication']['ActualEnergy'][0][0][:, 0]#Sonication.PlannedEnergy

                        ActualDuration = DATA_Mat['Sonication']['PlannedDuration'][0][0][:, 0]#Sonication.PlannedDuration
                        ActualEnergy = DATA_Mat['Sonication']['PlannedEnergy'][0][0][:, 0]
                        MaxAvgT = DATA_Mat['Sonication']['MaxAvgT'][0][0][:, 0]  # max average temperature which our target.

                        # sub_list = list(range(len(AvgPower)))
                        treatment_file = "./data/treatment_para/"
                        path = os.path.join(treatment_file, file) + ".csv"
                        data = pd.read_csv(path)
                        aa = data.iloc[:, 19]
                        # print(aa[1])
                        sub_list = aa[aa == 'No             '].index

                        AvgPower_in = []
                        Thickness_in = []
                        SDR_in = []
                        Angle_in=[]
                        ElementPosition_in=[]
                        # ActualEnergy_in=[]

                        MaxAvgT_out = []
                        for m in sub_list:
                            AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                            Thickness_in.append(Thickness[m][:, 0] * CT_density * OnOff[m][:, 0])
                            SDR_in.append(SDR[m][:, 0]* OnOff[m][:, 0])  # 是否受能量启动影像
                            Angle_in.append(Angle[m][:, 0]*OnOff[m][:,0])
                            ElementPosition_in.append(ElementPosition[m][:, :] * OnOff[m][:, 0][:, np.newaxis].
                                                      repeat(ElementPosition[m][:, :].shape[1], axis=1))
                        AvgPower_in = np.array(AvgPower_in)
                        Thickness_in = np.array(Thickness_in)/20000
                        SDR_in=np.array(SDR_in)
                        SDR_in=-np.nan_to_num(SDR_in)
                        Angle_in=np.array(Angle_in)/30
                        ElementPosition_in = np.array(ElementPosition_in)/200

                        ActualDuration_in = ActualDuration[sub_list]/10
                        ActualEnergy_in = ActualEnergy[sub_list]/10000
                        age_in = np.full_like (ActualDuration_in, 1)*age/100
                        dis_in = np.full_like (ActualDuration_in, 1)*dis/40

                        Brain_area_in=np.full_like (ActualDuration_in, 1)*Brain_area/400
                        Brain_size_in=np.full_like (ActualDuration_in, 1)*Brain_size/60
                        total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr
                        y_train = MaxAvgT[sub_list]
                        X_traina = AvgPower_in
                        X_traina=np.mean(X_traina, axis=1).reshape(-1, 1)
                        X_trainb=Thickness_in
                        X_trainb = np.mean(X_trainb, axis=1).reshape(-1, 1)

                        X_trainc=Angle_in#age_in[:, np.newaxis]
                        X_traind=dis_in[:, np.newaxis]
                        X_traine = ActualDuration_in[:, np.newaxis]
                        X_trainf=ActualEnergy_in[:, np.newaxis]

                        X_traing = Brain_area_in[:, np.newaxis]
                        X_trainh=Brain_size_in[:, np.newaxis]
                        X_traini = total_sdr_in[:, np.newaxis]
                        x_trainz=np.array(sub_list)[:, np.newaxis]/20


                        y_train = y_train[:, np.newaxis]

                        # 前向传播
                        # 将训练数据转换为 PyTorch 的 Tensor 类型
                        X_traina = torch.from_numpy(X_traina).float()
                        X_trainb = torch.from_numpy(X_trainb).float()
                        X_trainc = torch.from_numpy(X_trainc).float()
                        X_traind = torch.from_numpy(X_traind).float()
                        X_traine = torch.from_numpy(X_traine).float()
                        X_trainf = torch.from_numpy(X_trainf).float()
                        X_traing = torch.from_numpy(X_traing).float()
                        X_trainh = torch.from_numpy(X_trainh).float()
                        X_traini = torch.from_numpy(X_traini).float()
                        x_trainz = torch.from_numpy(x_trainz).float()

                        y_train = torch.from_numpy(y_train).float()
                        # X_m1_list = [X_traine, X_trainf, X_traini, X_traina, X_trainh, X_traing, X_traind, x_trainz]

                        X_m1_list = [X_trainc,X_traine, X_trainf, X_traini, X_traina, X_trainh, X_traing, X_traind, x_trainz]

                        outputs = model(X_m1_list)

                        loss = criterion(outputs, y_train)
                        # loss = criterion(outputs[:, np.newaxis], y_train).detach_().requires_grad_(True)
                        total_loss.append(loss)
                    mean_loss = np.mean(total_loss)
                return mean_loss

            def train(train_list, test_list):
                CT_value_file = "./data/CT_value_20231219"
                model.train()
                CT_value = {}
                for file in train_list:
                    ID = file.split('.')[0]
                    k = 0
                    path = os.path.join(CT_value_file, file)
                    with open(path, 'r', encoding='UTF-8') as f:
                        sub_data = []
                        for line in f.readlines():
                            k = k + 1
                            if k < 4:
                                continue
                            line = line.strip('\n')  # 去掉换行符\n
                            data_list = line.split('\t')[1:]
                            data_list = list(map(float, data_list))  # 将每一行以空格为分隔符转换成列表
                            mea = sum(data_list) / len(data_list)
                            sub_data.append(mea)  # 全部除以1000
                        sub_data = np.array(sub_data)
                        CT_value[ID] = sub_data

                # SDR score of patient;patients x 1
                path = "./data/CT_proj_data20231216.xlsx"
                look_up_table_row_start = 2
                look_up_table_row_number = 90
                skull_area={}
                skull_score={}
                measurement={}
                ages={}
                disea={}
                look_up_table_excel = load_workbook(path)
                look_up_table_all_sheet = look_up_table_excel.sheetnames
                look_up_table_sheet = look_up_table_excel[look_up_table_all_sheet[0]]
                for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
                    ID = look_up_table_sheet.cell(i, 1).value
                    sub_data1 = look_up_table_sheet.cell(i, 8).value
                    sub_data2 = look_up_table_sheet.cell(i, 11).value
                    sub_data3 = look_up_table_sheet.cell(i, 17).value
                    sub_data4 = look_up_table_sheet.cell(i, 20).value
                    sub_data5 = look_up_table_sheet.cell(i, 22).value
                    ages[ID]=sub_data1
                    disea[ID]=sub_data2
                    skull_area[ID] = sub_data3
                    skull_score[ID] = sub_data4
                    measurement[ID] = sub_data5

                Data_ver = "./data/Data_ver5"
                num_epochs = 10000
                best_val_loss = 1000  # initial
                best_epoch=1
                for epoch in range(num_epochs):
                    if epoch-best_epoch>512:
                        continue
                    random.shuffle(train_list)
                    # read .mat
                    for file in train_list:
                        CT_density = CT_value[file]
                        Brain_area=skull_area[file]
                        age=ages[file]
                        dis=disea[file]
                        total_sdr=skull_score[file]
                        Brain_size=measurement[file]
                        path = os.path.join(Data_ver, file)
                        DATA_Mat = scio.loadmat(path)
                        # print(DATA_Mat.keys())
                        AvgPower = DATA_Mat['Elements']['AvgPower'][0][0][0, :]
                        Thickness = DATA_Mat['Elements']['Thickness'][0][0][0, :]
                        SDR = DATA_Mat['Elements']['SDR'][0][0][0, :]
                        Angle = DATA_Mat['Elements']['Angle'][0][0][0, :]
                        ElementPosition = DATA_Mat['Elements']['ElementPosition'][0][0][0, :]
                        OnOff = DATA_Mat['Elements']['OnOff'][0][0][0, :]

                        # ActualDuration = DATA_Mat['Sonication']['ActualDuration'][0][0][:, 0]
                        # ActualEnergy = DATA_Mat['Sonication']['ActualEnergy'][0][0][:, 0]

                        ActualDuration = DATA_Mat['Sonication']['PlannedDuration'][0][0][:, 0]#Sonication.PlannedDuration
                        ActualEnergy = DATA_Mat['Sonication']['PlannedEnergy'][0][0][:, 0]
                        MaxAvgT = DATA_Mat['Sonication']['MaxAvgT'][0][0][:, 0]  # max average temperature which our target.
                        # num = random.randint(2 * int(len(AvgPower) / 3), len(AvgPower))
                        # sub_list = random.sample(range(len(AvgPower)), num)
                        # # sub_list = list(range(len(AvgPower)))
                        # random.shuffle(sub_list)
                        treatment_file = "./data/treatment_para/"
                        path = os.path.join(treatment_file, file) + ".csv"
                        data = pd.read_csv(path)
                        aa = data.iloc[:, 19]
                        # print(aa[1])
                        indices_of_no = aa[aa == 'No             '].index
                        indices_of_no=indices_of_no.tolist()
                        num = random.randint(2 * int(len(indices_of_no) / 3), len(indices_of_no))
                        sub_list = random.sample(indices_of_no, num)
                        # random.sample(indices_of_no, num)


                        AvgPower_in = []
                        Thickness_in = []
                        SDR_in = []
                        Angle_in=[]
                        ElementPosition_in=[]
                        # ActualEnergy_in=[]

                        MaxAvgT_out = []
                        for m in sub_list:
                            AvgPower_in.append(AvgPower[m][:, 0] * OnOff[m][:, 0])
                            Thickness_in.append(Thickness[m][:, 0] * CT_density * OnOff[m][:, 0])
                            SDR_in.append(SDR[m][:, 0]* OnOff[m][:, 0])  # 是否受能量启动影像
                            Angle_in.append(Angle[m][:, 0]*OnOff[m][:,0])
                            ElementPosition_in.append(ElementPosition[m][:, :] * OnOff[m][:, 0][:, np.newaxis].
                                                      repeat(ElementPosition[m][:, :].shape[1], axis=1))
                            # AvgPower_in=AvgPower[m][:,0]
                            # Thickness_in=Thickness[m][:,0]
                            # SDR_in=SDR[m][:,0]
                            # OnOff_in=OnOff[m][:,0]

                            # ActualDuration_in.append(ActualDuration[m][:, 0])
                            # ActualEnergy_in.append(ActualEnergy[m][:,0])
                            # MaxAvgT_out.append(MaxAvgT[m][:,0])
                        AvgPower_in = np.array(AvgPower_in)
                        Thickness_in = np.array(Thickness_in)/20000
                        SDR_in=np.array(SDR_in)
                        SDR_in=-np.nan_to_num(SDR_in)
                        Angle_in=np.array(Angle_in)/30
                        ElementPosition_in = np.array(ElementPosition_in)/200

                        ActualDuration_in = ActualDuration[sub_list]/10
                        ActualEnergy_in = ActualEnergy[sub_list]/10000
                        age_in = np.full_like (ActualDuration_in, 1)*age/100
                        dis_in = np.full_like (ActualDuration_in, 1)*dis/40
                        Brain_area_in=np.full_like (ActualDuration_in, 1)*Brain_area/400
                        Brain_size_in=np.full_like (ActualDuration_in, 1)*Brain_size/60
                        total_sdr_in = np.full_like(ActualDuration_in, 1) * total_sdr
                        y_train = MaxAvgT[sub_list]
                        if random.random()>0.5:
                            x,y=AvgPower_in.shape
                            AvgPower_in=AvgPower_in+np.random.rand(x,y)/10
                            Thickness_in=Thickness_in+np.random.rand(x,y)
                            SDR_in=SDR_in+np.random.rand(x,y)/10
                            Angle_in = Angle_in + np.random.rand(x, y) / 10
                            ElementPosition_in = ElementPosition_in + np.random.rand(x, y,3) /10

                            age_in = age_in + np.random.rand(x) / 10
                            dis_in = dis_in + np.random.rand(x) / 10

                            ActualDuration_in=ActualDuration_in+np.random.rand(x)/10
                            ActualEnergy_in=ActualEnergy_in+np.random.rand(x)/10

                            Brain_area_in=Brain_area_in+np.random.rand(x)/10
                            Brain_size_in=Brain_size_in+np.random.rand(x)/10
                            total_sdr_in=total_sdr_in++np.random.rand(x)/10

                            y_train=y_train+np.random.rand(x)
                        X_traina = AvgPower_in
                        X_traina=np.mean(X_traina, axis=1).reshape(-1, 1)
                        X_trainb=Thickness_in
                        X_trainb = np.mean(X_trainb, axis=1).reshape(-1, 1)
                        # X_trainc=SDR_in
                        # X_traind = np.concatenate((Angle_in[:, :, np.newaxis], ElementPosition_in),axis=2)
                        X_trainc=Angle_in#age_in[:, np.newaxis]
                        X_traind=dis_in[:, np.newaxis]

                        X_traine = ActualDuration_in[:, np.newaxis]
                        X_trainf=ActualEnergy_in[:, np.newaxis]
                        X_traing = Brain_area_in[:, np.newaxis]
                        X_trainh=Brain_size_in[:, np.newaxis]
                        X_traini = total_sdr_in[:, np.newaxis]
                        x_trainz=np.array(sub_list)[:, np.newaxis]/20

                        y_train = y_train[:, np.newaxis]

                        # 前向传播
                        # 将训练数据转换为 PyTorch 的 Tensor 类型
                        X_traina = torch.from_numpy(X_traina).float()#torch.tensor
                        X_trainb = torch.from_numpy(X_trainb).float()
                        X_trainc = torch.from_numpy(X_trainc).float()
                        X_traind = torch.from_numpy(X_traind).float()
                        X_traine = torch.from_numpy(X_traine).float()
                        X_trainf = torch.from_numpy(X_trainf).float()
                        X_traing = torch.from_numpy(X_traing).float()
                        X_trainh = torch.from_numpy(X_trainh).float()
                        X_traini = torch.from_numpy(X_traini).float()
                        x_trainz = torch.from_numpy(x_trainz).float()

                        y_train = torch.from_numpy(y_train).float()
                        # X_m1_list = [X_traine,X_trainf,X_traini,X_trainh,X_traing,X_traind,x_trainz]
                        # X_m1_list = [X_traine, X_trainf, X_traini, X_traina,X_trainh, X_traing, X_traind, x_trainz]
                        X_m1_list = [X_trainc,X_traine, X_trainf, X_traini, X_traina, X_trainh, X_traing, X_traind, x_trainz]
                        # X_train_list = [X_traina, X_trainb, X_trainc, X_traind, X_traine, X_trainf, X_traing, X_trainh,
                        #                 X_traini,x_trainz]
                        outputs = model(X_m1_list)
                        # y_train=sigmoid(y_train)
                        loss = criterion(outputs, y_train)#.detach_().requires_grad_(True)

                        # 反向传播和优化
                        optimizer.zero_grad()
                        loss.backward()
                        optimizer.step()

                    torch.save(model, latest_model)
                    val_loss = val_test(test_list)
                    if best_val_loss > val_loss:
                        best_epoch=epoch
                        best_val_loss = val_loss
                        print(f'Epoch [{epoch + 1}/{num_epochs}], Loss: {loss.item()}')
                        print(f'Epoch [{epoch + 1}/{num_epochs}], val_Loss: {val_loss.item()}')
                        print('Yayy! New best model saved!\n')
                        torch.save(model, best_model)
                        print('\n')

                    # if (epoch + 1) % 5 == 0:
                    #     print(f'Epoch [{epoch + 1}/{num_epochs}], Loss: {loss.item()}')
                    #     torch.save(model, latest_model)
                    #     val_loss = val_test(test_list)
                    #     print(f'Epoch [{epoch + 1}/{num_epochs}], val_Loss: {val_loss.item()}')
                    #     if best_val_loss > val_loss:
                    #         best_val_loss = val_loss
                    #         print('Yayy! New best model saved!\n')
                    #         torch.save(model, best_model)
                    #     print('\n')

            train(train_list,test_list)





